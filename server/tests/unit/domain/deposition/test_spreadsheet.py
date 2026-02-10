"""Unit tests for the Openpyxl spreadsheet adapter."""

from datetime import UTC, datetime
from io import BytesIO

from openpyxl import load_workbook

from osa.domain.semantics.model.schema import Schema
from osa.domain.semantics.model.value import (
    Cardinality,
    FieldDefinition,
    FieldType,
    TermConstraints,
)
from osa.domain.shared.model.srn import OntologySRN, SchemaSRN
from osa.infrastructure.persistence.adapter.spreadsheet import OpenpyxlSpreadsheetAdapter


def _make_schema_srn() -> SchemaSRN:
    return SchemaSRN.parse("urn:osa:localhost:schema:test@1.0.0")


def _make_ontology_srn() -> OntologySRN:
    return OntologySRN.parse("urn:osa:localhost:onto:sex@1.0.0")


def _make_text_field(name: str = "title", required: bool = True) -> FieldDefinition:
    return FieldDefinition(
        name=name,
        type=FieldType.TEXT,
        required=required,
        cardinality=Cardinality.EXACTLY_ONE,
        description="The title of the sample",
    )


def _make_term_field(
    name: str = "sex",
    onto_srn: OntologySRN | None = None,
) -> FieldDefinition:
    return FieldDefinition(
        name=name,
        type=FieldType.TERM,
        required=True,
        cardinality=Cardinality.EXACTLY_ONE,
        constraints=TermConstraints(ontology_srn=onto_srn or _make_ontology_srn()),
    )


def _make_schema(fields: list[FieldDefinition] | None = None) -> Schema:
    return Schema(
        srn=_make_schema_srn(),
        title="Test Schema",
        fields=fields or [_make_text_field()],
        created_at=datetime.now(UTC),
    )


class TestTemplateGeneration:
    def test_generates_valid_xlsx(self):
        adapter = OpenpyxlSpreadsheetAdapter()
        schema = _make_schema()
        content = adapter.generate_template(schema, {})
        wb = load_workbook(BytesIO(content))
        assert wb.active is not None

    def test_headers_match_field_names(self):
        adapter = OpenpyxlSpreadsheetAdapter()
        schema = _make_schema(
            fields=[_make_text_field("title"), _make_text_field("description", required=False)]
        )
        content = adapter.generate_template(schema, {})
        wb = load_workbook(BytesIO(content))
        ws = wb.active
        headers = [ws.cell(row=1, column=c).value for c in range(1, 3)]
        assert headers == ["title", "description"]

    def test_description_row_populated(self):
        adapter = OpenpyxlSpreadsheetAdapter()
        schema = _make_schema(fields=[_make_text_field("title")])
        content = adapter.generate_template(schema, {})
        wb = load_workbook(BytesIO(content))
        ws = wb.active
        desc = ws.cell(row=2, column=1).value
        assert desc == "The title of the sample"

    def test_required_field_header_bold(self):
        adapter = OpenpyxlSpreadsheetAdapter()
        schema = _make_schema(
            fields=[_make_text_field("title"), _make_text_field("optional_field", required=False)]
        )
        content = adapter.generate_template(schema, {})
        wb = load_workbook(BytesIO(content))
        ws = wb.active
        assert ws.cell(row=1, column=1).font.bold is True
        assert ws.cell(row=1, column=2).font.bold is not True

    def test_term_field_with_few_terms_gets_dropdown(self):
        adapter = OpenpyxlSpreadsheetAdapter()
        schema = _make_schema(fields=[_make_term_field()])
        onto_terms = {str(_make_ontology_srn()): ["male", "female", "other"]}
        content = adapter.generate_template(schema, onto_terms)
        wb = load_workbook(BytesIO(content))
        ws = wb.active
        # Check data validations exist
        assert len(ws.data_validations.dataValidation) > 0


class TestSpreadsheetParsing:
    def _create_upload(self, headers: list[str], rows: list[list]) -> bytes:
        """Create a valid spreadsheet upload with given headers and rows."""
        from openpyxl import Workbook

        wb = Workbook()
        ws = wb.active
        for col, header in enumerate(headers, 1):
            ws.cell(row=1, column=col, value=header)
        # Row 2 = descriptions (skip in parsing)
        for col in range(1, len(headers) + 1):
            ws.cell(row=2, column=col, value="description")
        # Row 3+ = data
        for row_idx, row_data in enumerate(rows, 3):
            for col, val in enumerate(row_data, 1):
                ws.cell(row=row_idx, column=col, value=val)

        buf = BytesIO()
        wb.save(buf)
        return buf.getvalue()

    def test_parse_valid_data(self):
        adapter = OpenpyxlSpreadsheetAdapter()
        schema = _make_schema(fields=[_make_text_field("title")])
        content = self._create_upload(["title"], [["Sample A"]])
        result = adapter.parse_upload(schema, content)
        assert result.metadata["title"] == "Sample A"
        assert result.errors == []

    def test_parse_missing_required_column(self):
        adapter = OpenpyxlSpreadsheetAdapter()
        schema = _make_schema(fields=[_make_text_field("title"), _make_text_field("organism")])
        content = self._create_upload(["title"], [["Sample A"]])
        result = adapter.parse_upload(schema, content)
        assert len(result.errors) > 0
        assert any("organism" in e.field for e in result.errors)

    def test_parse_ignores_unrecognized_columns(self):
        adapter = OpenpyxlSpreadsheetAdapter()
        schema = _make_schema(fields=[_make_text_field("title")])
        content = self._create_upload(["title", "extra_col"], [["Sample A", "ignored"]])
        result = adapter.parse_upload(schema, content)
        assert "title" in result.metadata
        assert "extra_col" not in result.metadata

    def test_parse_unrecognized_columns_generate_warnings(self):
        adapter = OpenpyxlSpreadsheetAdapter()
        schema = _make_schema(fields=[_make_text_field("title")])
        content = self._create_upload(["title", "extra_col"], [["Sample A", "ignored"]])
        result = adapter.parse_upload(schema, content)
        assert any("extra_col" in w for w in result.warnings)

    def test_parse_missing_required_value(self):
        adapter = OpenpyxlSpreadsheetAdapter()
        schema = _make_schema(fields=[_make_text_field("title")])
        content = self._create_upload(["title"], [[None]])
        result = adapter.parse_upload(schema, content)
        assert len(result.errors) > 0

    def test_parse_optional_field_allows_empty(self):
        adapter = OpenpyxlSpreadsheetAdapter()
        schema = _make_schema(
            fields=[
                _make_text_field("title"),
                _make_text_field("notes", required=False),
            ]
        )
        content = self._create_upload(["title", "notes"], [["Sample A", None]])
        result = adapter.parse_upload(schema, content)
        assert result.errors == []
