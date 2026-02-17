"""Openpyxl-based spreadsheet adapter for template generation and parsing."""

from io import BytesIO
from typing import Any

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.worksheet.datavalidation import DataValidation

from osa.domain.deposition.port.spreadsheet import (
    SpreadsheetError,
    SpreadsheetParseResult,
    SpreadsheetPort,
)
from osa.domain.semantics.model.schema import Schema
from osa.domain.semantics.model.value import FieldDefinition, FieldType

# Ontologies with <=20 terms get dropdown validation; others get an instruction note.
_MAX_DROPDOWN_TERMS = 20
_REQUIRED_FILL = PatternFill(start_color="FFFFEE", end_color="FFFFEE", fill_type="solid")
_REQUIRED_FONT = Font(bold=True)
_DESC_FONT = Font(italic=True, color="888888")


class OpenpyxlSpreadsheetAdapter(SpreadsheetPort):
    def generate_template(
        self,
        schema: Schema,
        ontology_terms_by_srn: dict[str, list[str]],
    ) -> bytes:
        wb = Workbook()
        ws = wb.active
        ws.title = "Data"

        for col_idx, field in enumerate(schema.fields, 1):
            # Row 1: headers
            cell = ws.cell(row=1, column=col_idx, value=field.name)
            if field.required:
                cell.font = _REQUIRED_FONT
                cell.fill = _REQUIRED_FILL

            # Row 2: descriptions
            desc_cell = ws.cell(row=2, column=col_idx, value=field.description or "")
            desc_cell.font = _DESC_FONT

            # Add dropdown for term fields with small ontologies
            if field.type == FieldType.TERM and field.constraints:
                onto_srn_str = str(field.constraints.ontology_srn)
                terms = ontology_terms_by_srn.get(onto_srn_str, [])
                if terms and len(terms) <= _MAX_DROPDOWN_TERMS:
                    formula = '"' + ",".join(terms) + '"'
                    dv = DataValidation(
                        type="list", formula1=formula, allow_blank=not field.required
                    )
                    dv.sqref = f"{ws.cell(row=3, column=col_idx).coordinate}:{ws.cell(row=1000, column=col_idx).coordinate}"
                    ws.add_data_validation(dv)
                elif terms:
                    # Too many terms — add an instruction note in description
                    ws.cell(
                        row=2,
                        column=col_idx,
                        value=f"Select from ontology {onto_srn_str} ({len(terms)} terms)",
                    )

        # Auto-size columns
        for col_idx in range(1, len(schema.fields) + 1):
            ws.column_dimensions[ws.cell(row=1, column=col_idx).column_letter].width = 20

        buf = BytesIO()
        wb.save(buf)
        return buf.getvalue()

    def parse_upload(
        self,
        schema: Schema,
        content: bytes,
    ) -> SpreadsheetParseResult:
        wb = load_workbook(BytesIO(content), data_only=True)
        ws = wb.active

        errors: list[SpreadsheetError] = []
        warnings: list[str] = []
        metadata: dict[str, Any] = {}

        # Read headers from row 1
        headers: list[str | None] = []
        for col_idx in range(1, ws.max_column + 1):
            val = ws.cell(row=1, column=col_idx).value
            headers.append(str(val) if val is not None else None)

        # Build field lookup
        field_by_name: dict[str, FieldDefinition] = {f.name: f for f in schema.fields}

        # Check for missing required columns
        present_names = {h for h in headers if h is not None}
        for field in schema.fields:
            if field.required and field.name not in present_names:
                errors.append(
                    SpreadsheetError(
                        field=field.name,
                        message=f"Required column '{field.name}' is missing",
                    )
                )

        # Warn about unrecognized columns
        for h in headers:
            if h is not None and h not in field_by_name:
                warnings.append(f"Unrecognized column '{h}' will be ignored")

        # Parse data row (row 3 — row 2 is descriptions)
        data_row = 3
        for col_idx, header in enumerate(headers, 1):
            if header is None or header not in field_by_name:
                continue
            field = field_by_name[header]
            value = ws.cell(row=data_row, column=col_idx).value

            if value is None or (isinstance(value, str) and value.strip() == ""):
                if field.required:
                    errors.append(
                        SpreadsheetError(
                            field=field.name,
                            message=f"Required field '{field.name}' is empty",
                        )
                    )
                continue

            metadata[field.name] = value

        return SpreadsheetParseResult(
            metadata=metadata,
            warnings=warnings,
            errors=errors,
        )
